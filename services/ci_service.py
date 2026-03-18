from writers.excel_writer import processar_planilha_gas
from writers.excel_writer_oleo import processar_planilha_oleo
from openpyxl import load_workbook

def executar_fluxo(ci_path: str, dados: dict, tipo: str):

    if tipo == "gas":

        processar_planilha_gas(
            caminho_excel=ci_path,
            dados=dados
        )

    elif tipo == "oleo":
        processar_planilha_oleo(
            caminho_excel=ci_path,
            dados=dados
        )

    else:
        raise ValueError("Tipo de CI não suportado")



def identificar_tipo_ci(caminho_excel):

    wb = load_workbook(caminho_excel, read_only=True)

    abas = set(wb.sheetnames)

    wb.close()

    abas_gas = {
        "Report",
        "Gas parameters",
        "Meter run parameters",
        "Coef Disc & Expansao"
    }

    abas_oleo = {
        "Report",
        "Meter run parameters",
        "Gráfico - Variação Linear"
    }

    if abas_gas.issubset(abas):
        return "gas"
    elif abas_oleo.issubset(abas):
        return "oleo"

    return "nao_suportado"